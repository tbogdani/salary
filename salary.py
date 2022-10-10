from openpyxl import load_workbook

fn = 'st.xlsx'
fn0 = 'data.xlsx'

wb0 = load_workbook(filename=fn0, data_only=True)
ws0 = wb0['data']
wb = load_workbook(fn)
ws = wb['data']


def pr_grade(q, grade):
    if q < 80:
        pr = 0
    elif q < 90:
        pr = ws.cell(row=5, column=22 + grade).value
    elif q < 100:
        pr = ws.cell(row=6, column=22 + grade).value
    elif q < 120:
        pr = ws.cell(row=7, column=22 + grade).value + ws.cell(row=8, column=22 + grade).value * (q - 100)
    else:
        pr = ws.cell(row=9, column=22 + grade).value
    return pr


grades = ['Оператор', 'Специалист', 'Старший', 'Ведущий']
z0 = 40000 / int(ws.cell(row=2, column=4).value)

for i in range(6, 21):
    ws.cell(row=i, column=5).value = ws0.cell(row=i * 2 + 8, column=2).value
    ws0.cell(row=i * 2 + 9, column=5).value = ws0.cell(row=i * 2 + 9, column=4).value / ws0.cell(row=i * 2 + 9,
                                                                                                 column=3).value * 100
    ws.cell(row=i, column=6).value = round(ws0.cell(row=i * 2 + 9, column=5).value)

    z = z0 * int(ws.cell(row=i, column=5).value)
    k = float(ws.cell(row=i, column=6).value)

    prem = round(float(pr_grade(k, grades.index(ws.cell(row=i, column=4).value))), 2)
    z = round(z, 2)

    ws.cell(row=i, column=7, value=z)
    ws.cell(row=i, column=8, value=prem)
    ws.cell(row=i, column=9, value=z + prem)

wb0.save(fn0)
wb0.close()

wb.save(fn)
wb.close()
